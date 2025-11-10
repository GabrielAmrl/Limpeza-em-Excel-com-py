#!/usr/bin/env python3
import os
import unicodedata
import pandas as pd
from pandas.tseries.offsets import BDay
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BUSINESS_DAYS_BACK = 3
DATE_COLUMNS_TO_SPLIT = ["Data de abertura", "Resposta Área"]
COL_CANAL_ORIGEM = "Canal de Origem"
COL_STATUS = "Status"
COL_ASSUNTO = "Assunto"
COL_PROCESSO_ORIGEM = "Processo Origem"
COL_SEGMENTO = "Segmento"

def normalize_str(s):
    """Remove acentuação, normaliza espaços e coloca em maiúsculas."""
    s = "" if s is None else str(s)
    s = s.strip()
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("utf-8")
    return " ".join(s.split()).upper()

def text_to_column_tab(df, col):
    if col in df.columns:
        df[col] = df[col].astype(str).fillna("").apply(lambda s: s.split("\t")[0].strip())
    return df

def drop_columns_by_letter(df, letters):
    cols = list(df.columns)
    drop_idxs = []
    for letter in letters:
        idx = ord(letter.upper()) - ord('A')
        if 0 <= idx < len(cols):
            drop_idxs.append(idx)
    drop_names = [cols[i] for i in sorted(set(drop_idxs), reverse=True)]
    return df.drop(columns=drop_names, errors='ignore')

def drop_total_row(df):
    mask_total = df.apply(lambda r: r.astype(str).str.contains("total", case=False, na=False)).any(axis=1)
    return df.loc[~mask_total].reset_index(drop=True)

def parse_dates(df, col):
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce')
    return df

def append_and_tag(df_out, rows, new_status):
    if rows is None or rows.empty:
        return df_out
    tmp = rows.copy()
    tmp[COL_STATUS] = new_status
    return pd.concat([df_out, tmp], ignore_index=True, sort=False)

def process_dataframe(df):
    df = drop_columns_by_letter(df, ["C", "A"])
    df = drop_total_row(df)

    for col in DATE_COLUMNS_TO_SPLIT:
        df = text_to_column_tab(df, col)
        df = parse_dates(df, col)

    # === Cálculo correto do corte de 3 dias úteis atrás ===
    today = pd.Timestamp.today().normalize()
    cutoff = (today - BDay(BUSINESS_DAYS_BACK)).normalize()
    # Exemplo: hoje 22/10/2025 -> cutoff = 17/10/2025
    # IMPORTANT: não aplicamos filtro global que remova linhas; deixamos os testes usarem <= cutoff

    df_out = pd.DataFrame(columns=df.columns)

    # Em processo acima de 48h (Data de abertura <= cutoff)
    cond_em_processo = (
        df[COL_STATUS].astype(str).str.strip().str.lower() == "em processo"
    ) & (
        df["Data de abertura"].notna()
    ) & (
        df["Data de abertura"] <= cutoff
    )
    df_out = append_and_tag(df_out, df.loc[cond_em_processo], "Em processo acima de 48hrs")

    # Respondido acima de 48h (Resposta Área <= cutoff)
    cond_respondido_with_date = (
        df[COL_STATUS].astype(str).str.strip().str.lower() == "respondido"
    ) & (
        df["Resposta Área"].notna()
    ) & (
        df["Resposta Área"] <= cutoff
    )
    df_out = append_and_tag(df_out, df.loc[cond_respondido_with_date], "Respondido acima de 48hrs")

    # Respondido campo vazio
    cond_respondido_empty = (
        df[COL_STATUS].astype(str).str.strip().str.lower() == "respondido"
    ) & (
        df["Resposta Área"].isna() | (df["Resposta Área"].astype(str).str.strip() == "")
    )
    vazio = df.loc[cond_respondido_empty].copy()
    vazio["Resposta Área"] = vazio["Resposta Área"].fillna("").replace("", "Respondido - campo vazio")
    df_out = append_and_tag(df_out, vazio, "Respondido - campo vazio")

    # Em verificação - campo vazio
    if "Prazo reclamação" in df.columns:
        cond_verif_empty = (
            df[COL_STATUS].astype(str).str.strip().str.lower() == "em verificação"
        ) & (
            df["Prazo reclamação"].isna() | (df["Prazo reclamação"].astype(str).str.strip() == "")
        )
        df_out = append_and_tag(df_out, df.loc[cond_verif_empty], "Em verificação - campo vazio")

    # Remover Solicitação em Tipo do caso
    df_out.columns = (
        df_out.columns.str.strip()
        .str.normalize('NFKD')
        .str.encode('ascii', errors='ignore')
        .str.decode('utf-8')
        .str.lower()
    )
    tipo_col = next((col for col in df_out.columns if "tipo" in col and "caso" in col), None)
    if tipo_col:
        df_out = df_out[
            ~(
                df_out[tipo_col]
                .astype(str)
                .str.strip()
                .str.normalize('NFKD')
                .str.encode('ascii', errors='ignore')
                .str.decode('utf-8')
                .str.lower()
                .eq("solicitacao")
            )
        ].reset_index(drop=True)

    # Remover Canal de Origem = OUV ORGAO DEFESA - PROCON INTIMACAO
    canal_col = next((col for col in df_out.columns if "canal" in col and "origem" in col), None)
    if canal_col:
        df_out = df_out[
            ~(
                df_out[canal_col]
                .astype(str)
                .str.strip()
                .str.upper()
                .eq("OUV ORGAO DEFESA - PROCON INTIMACAO")
            )
        ].reset_index(drop=True)

    # Destacar Respondido - campo vazio com assunto específico
    assunto_bad = {"RECURSO DE TOI OUV", "RECURSO DE RESSARCIMENTO"}
    df_out["__HIGHLIGHT__"] = False
    if COL_ASSUNTO.lower() in df_out.columns and COL_STATUS.lower() in df_out.columns:
        mask_red = (
            df_out[COL_ASSUNTO.lower()].astype(str).str.strip().str.upper().isin(assunto_bad)
            & df_out[COL_STATUS.lower()].astype(str).str.strip().str.lower().eq("respondido - campo vazio")
        )
        df_out.loc[mask_red, "__HIGHLIGHT__"] = True

    # === Regra 13 - Mapeamento de Processo Origem (exatas / com normalização onde útil) ===
    if COL_PROCESSO_ORIGEM.lower() in df_out.columns:
        def map_processo(row):
            po_raw = row.get(COL_PROCESSO_ORIGEM.lower(), "")
            canal_raw = row.get(COL_CANAL_ORIGEM.lower(), "")
            seg_raw = row.get(COL_SEGMENTO.lower(), "")

            po = normalize_str(po_raw)
            canal = normalize_str(canal_raw)
            seg = normalize_str(seg_raw)

            # ANEEL + Segmento 2º ou 3º NÍVEL ANEEL
            if po == "ANEEL":
                if "2" in seg and "NIVEL" in seg and "ANEEL" in seg:
                    return "ANEEL 2°N"
                if "3" in seg and "NIVEL" in seg and "ANEEL" in seg:
                    return "ANEEL 3°N"

            # ORGÃO DEFESA CONSUMIDOR + Canal OUVIDORIA - CONSUMIDOR GOV -> GOV
            if po == "ORGAO DEFESA CONSUMIDOR":
                # verificar substring para pegar variações possíveis
                if "CONSUMIDOR GOV" in canal or canal == "OUVIDORIA CONSUMIDOR GOV":
                    return "GOV"
                # Procon variants -> ODC
                procon_variants = {
                    "OUV ORGAO DEFESA - PROCON ASSEMBLEIA",
                    "OUV ORGAO DEFESA - PROCON CIP",
                    "OUV ORGAO DEFESA - PROCON CIP ELETRONICA",
                    "OUV ORGAO DEFESA - PROCON NOTIFICACAO",
                    "OUV ORGAO DEFESA - PROCON TEL",
                    "OUVIDORIA - CARTA/OFICIO"
                }
                if any(v in canal for v in procon_variants) or "PROCON" in canal:
                    return "ODC"

            # ATENDIMENTO OUVIDORIA com canais específicos
            if po == "ATENDIMENTO OUVIDORIA":
                if "CALL CENTER" in canal:
                    return "BKO"
                if "E MAIL" in canal or "EMAIL" in canal:
                    return "EMAIL"
                if "AREA INTERNA" in canal:
                    return "ÁREA INTERNA"
                if "CONSELHO CONSUMIDOR" in canal:
                    return "CONSELHO"

            # preserva o valor original se nenhuma regra aplicou
            return po_raw

        df_out[COL_PROCESSO_ORIGEM.lower()] = df_out.apply(map_processo, axis=1)

    return df_out

def save_with_formatting(df_out, original_path, out_path=None):
    if out_path is None:
        base, ext = os.path.splitext(original_path)
        out_path = base + "_transformado.xlsx"
    df_out.to_excel(out_path, index=False, engine='openpyxl')

    wb = load_workbook(out_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        highlight_idx = header.index("__HIGHLIGHT__") + 1
    except ValueError:
        highlight_idx = None

    if highlight_idx:
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=highlight_idx).value).strip().lower() == "true":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = red_fill
        ws.delete_cols(highlight_idx)

    wb.save(out_path)
    return out_path

# --- INTERFACE ---
class App:
    def __init__(self, root):
        self.root = root
        root.title("Transformador de Planilha")
        root.geometry("600x200")
        self.filepath = None

        tk.Label(root, text="Selecione um arquivo CSV ou XLSX para transformar:").pack(pady=6)
        tk.Button(root, text="Abrir arquivo", command=self.open_file).pack(pady=4)
        self.lbl_file = tk.Label(root, text="Nenhum arquivo selecionado")
        self.lbl_file.pack(pady=4)
        tk.Button(root, text="Executar transformação", command=self.run_transform).pack(pady=8)

    def open_file(self):
        path = filedialog.askopenfilename(filetypes=[("Planilhas", "*.csv *.xlsx *.xls")])
        if path:
            self.filepath = path
            self.lbl_file.config(text=os.path.basename(path))

    def run_transform(self):
        if not self.filepath:
            messagebox.showerror("Erro", "Nenhum arquivo selecionado")
            return
        try:
            df = pd.read_excel(self.filepath, engine='openpyxl', header=15)
            df_out = process_dataframe(df)
            out_path = save_with_formatting(df_out, self.filepath)
            messagebox.showinfo("Sucesso", f"Arquivo salvo em:\n{out_path}")
        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")

def main():
    root = tk.Tk()
    App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
